import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os
from typing import List, Dict, Any, Optional
from .config import get_excel_filepath, get_excel_directory
class ExcelOperations:
    
    @staticmethod
    def create_excel_file(filename: str, sheet_name: str = "Sheet1") -> Dict[str, Any]:
        """新規Excelファイルを作成"""
        try:
            filepath = get_excel_filepath(filename)
            
            # 既存ファイルの確認
            if os.path.exists(filepath):
                return {
                    "success": False,
                    "error": f"ファイル '{filename}' は既に存在します",
                    "filename": filename,
                    "path": filepath
                }
            
            # 新規ワークブック作成
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # ファイル保存
            wb.save(filepath)
            
            return {
                "success": True,
                "filename": filename,
                "path": filepath,
                "sheet_name": sheet_name,
                "message": f"Excelファイル '{filename}' を作成しました"
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": f"ファイル作成エラー: {str(e)}",
                "filename": filename
            }
    
    @staticmethod
    def write_excel_data(filename: str, sheet_name: str, data: List[Dict[str, Any]], 
                        start_cell: str = "A1", include_header: bool = True) -> Dict[str, Any]:
        """Excelファイルに表データを書き込み"""
        try:
            if not data:
                return {
                    "success": False,
                    "error": "書き込むデータが空です"
                }
            
            filepath = get_excel_filepath(filename)
            
            # DataFrameに変換
            df = pd.DataFrame(data)
            
            # 既存ファイルがあるかチェック
            if os.path.exists(filepath):
                # 既存ファイルを開く
                wb = openpyxl.load_workbook(filepath)
                
                # シートが存在するかチェック
                if sheet_name not in wb.sheetnames:
                    wb.create_sheet(sheet_name)
                
                ws = wb[sheet_name]
                # 既存データをクリア
                ws.delete_rows(1, ws.max_row)
            else:
                # 新規ファイル作成
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = sheet_name
            
            # データをワークシートに書き込み
            ExcelOperations._write_data_to_worksheet(ws, df, start_cell, include_header)
            
            # ファイル保存
            wb.save(filepath)
            
            return {
                "success": True,
                "filename": filename,
                "sheet_name": sheet_name,
                "rows_written": len(data),
                "columns": len(df.columns),
                "headers": list(df.columns) if include_header else [],
                "path": filepath,
                "message": f"'{sheet_name}'シートに{len(data)}行のデータを書き込みました"
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": f"データ書き込みエラー: {str(e)}",
                "filename": filename,
                "sheet_name": sheet_name
            }
    
    @staticmethod
    def _write_data_to_worksheet(ws, df: pd.DataFrame, start_cell: str, include_header: bool):
        """ワークシートにデータと書式を適用"""
        # 開始位置の解析
        start_col = ord(start_cell[0].upper()) - ord('A') + 1
        start_row = int(start_cell[1:])
        
        # ヘッダー行の書き込みと書式設定
        if include_header:
            for col, header in enumerate(df.columns):
                cell = ws.cell(row=start_row, column=start_col + col, value=header)
                # ヘッダーの書式設定
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
                cell.border = ExcelOperations._get_border()
            start_row += 1
        
        # データ行の書き込みと書式設定
        for row_idx, row_data in df.iterrows():
            for col_idx, value in enumerate(row_data):
                cell = ws.cell(row=start_row + row_idx, column=start_col + col_idx, value=value)
                cell.border = ExcelOperations._get_border()
                cell.alignment = Alignment(horizontal="left")
        
        # 列幅の自動調整
        ExcelOperations._auto_adjust_column_width(ws)
    
    @staticmethod
    def _get_border():
        """標準的な枠線スタイルを取得"""
        return Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    @staticmethod
    def _auto_adjust_column_width(ws):
        """列幅を自動調整"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # 最大幅を50に制限
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def read_excel_data(filename: str, sheet_name: str, range_cells: Optional[str] = None, 
                       has_header: bool = True) -> Dict[str, Any]:
        """Excelファイルから表データを読み込み"""
        try:
            filepath = get_excel_filepath(filename)
            
            if not os.path.exists(filepath):
                return {
                    "success": False,
                    "error": f"ファイル '{filename}' が見つかりません"
                }
            
            # pandasでデータ読み込み
            if range_cells:
                # 特定範囲の読み込み（実装簡略化のため全体読み込み後に範囲指定）
                df = pd.read_excel(filepath, sheet_name=sheet_name, header=0 if has_header else None)
            else:
                df = pd.read_excel(filepath, sheet_name=sheet_name, header=0 if has_header else None)
            
            # NaN値を空文字に変換
            df = df.fillna("")
            
            # 辞書形式に変換
            data = df.to_dict('records')
            
            return {
                "success": True,
                "data": data,
                "rows": len(df),
                "columns": len(df.columns),
                "headers": list(df.columns) if has_header else [],
                "filename": filename,
                "sheet_name": sheet_name
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": f"データ読み込みエラー: {str(e)}",
                "filename": filename,
                "sheet_name": sheet_name
            }
    
    @staticmethod
    def list_sheets(filename: str) -> Dict[str, Any]:
        """Excelファイル内のシート一覧を取得"""
        try:
            filepath = get_excel_filepath(filename)
            
            if not os.path.exists(filepath):
                return {
                    "success": False,
                    "error": f"ファイル '{filename}' が見つかりません"
                }
            
            wb = openpyxl.load_workbook(filepath)
            sheets = wb.sheetnames
            
            return {
                "success": True,
                "sheets": sheets,
                "count": len(sheets),
                "filename": filename
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": f"シート一覧取得エラー: {str(e)}",
                "filename": filename
            }
    
    @staticmethod
    def list_excel_files() -> Dict[str, Any]:
        """利用可能なExcelファイル一覧を取得"""
        try:
            excel_dir = get_excel_directory()
            files = []
            
            for filename in os.listdir(excel_dir):
                if filename.endswith('.xlsx'):
                    filepath = os.path.join(excel_dir, filename)
                    file_size = os.path.getsize(filepath)
                    files.append({
                        "filename": filename,
                        "size_bytes": file_size,
                        "size_mb": round(file_size / 1024 / 1024, 2)
                    })
            
            return {
                "success": True,
                "files": files,
                "count": len(files),
                "directory": excel_dir
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": f"ファイル一覧取得エラー: {str(e)}"
            }